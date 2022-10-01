import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inventory = {}


# Task1 Calculate how many products per supplier
for product_row in range(2, product_list.max_row + 1):  # row 2 to 76 will run from 2 to 75
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row,2).value
    price = product_list.cell(product_row,3).value
    product_number = product_list.cell(product_row,1).value
    # for writing into another column
    inventory_price = product_list.cell(product_row,5)

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

# Task2 calculation total value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory*price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# Task 3 calculation of products with inventory less than 10
    if inventory<10:
        products_under_10_inventory[int(product_number)] = int(inventory)


# Task 4 Add value for total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inventory)

inv_file.save("inventory_with_total_value.xlsx")
